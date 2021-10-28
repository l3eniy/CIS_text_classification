import openpyxl

# Folgende Files werden benötigt:
# --> CIS_CAT_RESULTS.xlsx
# --> CIS_FUN_MAPPED.xlsx


SPALTE_CIS = 12
ZEILE_ERSTES_CIS = 15
ERSTES_FUN_REQUIREMENT = 19
CIS_CAT_REQ_SPALTE = 4

# CIS CAT Ergebnisse laden
CIS_CAT_File = openpyxl.load_workbook("CIS_CAT_RESULTS.xlsx")
CIS_CAT_Sheet = CIS_CAT_File['Export XLS - 0']


def get_cis_cat_result_from_cis_req(CIS_Req_String):
    '''
    Funktion die CIS CAT Ergebnis zurückgibt, wenn sie eins findet, oder None, wenn sie keins findet
    :param CIS_Req_String: das CIS Requirement, nach dem gesucht wird
    :return: CIS CAT ergebnis als String oder None
    !!! greift auf geöffnetes CIS_CAT_Sheet zu !!!
    '''
    for row_ciscat in CIS_CAT_Sheet.iter_rows(min_row=2, min_col=CIS_CAT_REQ_SPALTE, max_col=CIS_CAT_REQ_SPALTE):
        cell_ciscat = row_ciscat[0]
        # print("[+] "+ cell_ciscat.value + " == " + CIS_Req_String)
        if cell_ciscat.value == CIS_Req_String:
            # print("Success!")
            return CIS_CAT_Sheet["H" + str(row_ciscat[0].row)].value
    print("[+] " + str(CIS_Req_String) + "  KEINE ZUORDNUNG in CIS CAT Ergebnis GEFUNDEN!")
    return None


# Mapping Excel Blatt öffnen
theFile = openpyxl.load_workbook('CIS_FUN_MAPPED.xlsx')
currentSheet = theFile['RHEL 7 BM v.2.2.0_gesamt']

# Über jede Zeile (CIS Req) im Mapping Excel Blatt gehen
for row in currentSheet.iter_rows(min_row=ZEILE_ERSTES_CIS, min_col=SPALTE_CIS, max_col=SPALTE_CIS):
    for cell in row:
        # abgleich machen, ob CIS Requirement auch in CIS Cat vorkommt, wenn nicht --> nächstes CIS Req:
        if get_cis_cat_result_from_cis_req(cell.value) is None:
            continue

        # durch jede Spalte iterieren
        for row_ in currentSheet.iter_rows(max_row=row[0].row, min_row=row[0].row, min_col=ERSTES_FUN_REQUIREMENT):
            for cell_ in row_:

                # Wenn Must oder Should in cell_.value, dann das entsprechende Ergebnis aus CIS CAT reinschreiben
                if cell_.value == "Must":
                    cell_.value = "Must: " + get_cis_cat_result_from_cis_req(cell.value)
                if cell_.value == "Should":
                    cell_.value = "Should: " + get_cis_cat_result_from_cis_req(cell.value)
theFile.save('CIS_FUN_MAPPED_processed.xlsx')
theFile.close()
CIS_CAT_File.close()
